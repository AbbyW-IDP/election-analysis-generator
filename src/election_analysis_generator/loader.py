"""
loader.py
---------
Election data loaders.

Two concrete classes handle different input file formats:

  LoadSummary       – reads a county-wide summary CSV (one row per
                      candidate per contest) and populates the candidates
                      table. This is what sync() orchestrates.

  LoadPrecinctDetail – reads the precinct-detail Excel workbook (one sheet
                       per party per contest, all precincts in that sheet)
                       and populates candidate_precinct_results.

Both classes are idempotent: the database's loaded_files registry
tracks which filenames have been processed, so running sync() or
load_detail_excel() a second time with the same file is a safe no-op.

Elections are configured in elections.toml. Each entry maps a CSV (and
optionally a detail Excel file) to election metadata. The loader reads
the toml, checks which sources have not been loaded yet, and loads any new
ones.
"""

from __future__ import annotations

import re
import tomllib
from datetime import date
from pathlib import Path

import pandas as pd

from .db import ElectionDatabase
from .models import Election
from .normalize import normalize_contest_name

DEFAULT_SOURCES_DIR = Path("sources")
DEFAULT_CONFIG_PATH = Path("elections.toml")

VALID_CATEGORIES = frozenset(
    {"Consolidated", "Consolidated Primary", "General", "General Primary"}
)
VALID_ELECTION_TYPES = frozenset({"presidential", "midterm"})

# Required columns (post-normalisation names).
REQUIRED_CSV_COLUMNS = frozenset({"contest_name_raw", "party", "total_votes"})

# All known optional columns. Any absent ones are added as NaN so the rest
# of the pipeline doesn't need to guard against missing columns.
OPTIONAL_CSV_COLUMNS = (
    "line_number",
    "choice_name",
    "percent_of_votes",
    "registered_voters",
    "ballots_cast",
    "num_precinct_total",
    "num_precinct_rptg",
    "over_votes",
    "under_votes",
)

# Candidate name cell values that indicate no real candidate was filed.
_NO_CANDIDATE_MARKERS = frozenset(
    {"NO CANDIDATE/CANDIDATO", "NO CANDIDATE", "CANDIDATO"}
)


# ---------------------------------------------------------------------------
# Module-level helpers (shared between subclasses)
# ---------------------------------------------------------------------------


def _normalize_csv_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize CSV column names to internal conventions."""
    df = df.copy()
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    return df.rename(
        columns={
            "contest_name": "contest_name_raw",
            "party_name": "party",
        }
    )


def _validate_csv_columns(df: pd.DataFrame, path: Path) -> pd.DataFrame:
    """
    Check that all required columns are present and add any missing optional
    columns as NaN.

    Raises ValueError naming every missing required column.
    """
    missing_required = REQUIRED_CSV_COLUMNS - set(df.columns)
    if missing_required:
        display = {
            "contest_name_raw": "contest name",
            "party": "party",
            "total_votes": "total votes",
        }
        friendly = sorted(display.get(c, c) for c in missing_required)
        raise ValueError(
            f"{path.name} is missing required column(s): {', '.join(friendly)}"
        )

    for col in OPTIONAL_CSV_COLUMNS:
        if col not in df.columns:
            df[col] = None

    return df


def _year_from_filename(filename: str) -> int | None:
    """
    Extract the election year from a filename.
    Tries the start of the stem first (handles date-suffixed names like
    2022-general-primary-2022-07-19.csv), then falls back to first 20xx match.
    """
    stem = Path(filename).stem
    match = re.match(r"(20\d{2})", stem)
    if match:
        return int(match.group(1))
    match = re.search(r"(20\d{2})", stem)
    return int(match.group(1)) if match else None


def _validate_config_entry(entry: dict, key: str) -> None:
    """Raise ValueError if a config entry has invalid category or election_type."""
    category = entry.get("category")
    election_type = entry.get("election_type")

    if category is not None and category not in VALID_CATEGORIES:
        raise ValueError(
            f"[elections.{key}] Invalid category {category!r}. "
            f"Must be one of: {sorted(VALID_CATEGORIES)}"
        )
    if election_type is not None and election_type not in VALID_ELECTION_TYPES:
        raise ValueError(
            f"[elections.{key}] Invalid election_type {election_type!r}. "
            f"Must be one of: {sorted(VALID_ELECTION_TYPES)}"
        )


def load_elections_config(config_path: Path = DEFAULT_CONFIG_PATH) -> list[dict]:
    """
    Read elections.toml and return a list of election config dicts.

    Each entry must have: name, summary_file.
    Optional: year, election_date, results_last_updated, category,
              election_type, ballots_cast, registered_voters, detail_file.

    Raises ValueError if any entry has an invalid category or election_type.
    """
    if not config_path.exists():
        return []
    with open(config_path, "rb") as f:
        raw = tomllib.load(f)

    entries = []
    for key, cfg in raw.get("elections", {}).items():
        _validate_config_entry(cfg, key)
        entries.append(cfg)
    return entries


# ---------------------------------------------------------------------------
# Abstract base
# ---------------------------------------------------------------------------


class _LoaderBase:
    """
    Shared constructor for all loader classes.

    Both LoadSummary and LoadPrecinctDetail take a single ElectionDatabase
    argument and store it as self._db.
    """

    def __init__(self, db: ElectionDatabase) -> None:
        self._db = db


# ---------------------------------------------------------------------------
# Subclass 1: summary CSV loader
# ---------------------------------------------------------------------------


class LoadSummary(_LoaderBase):
    """
    Reads county-wide summary CSVs and populates the candidates table.

    Call sync() to automatically load any elections defined in elections.toml
    whose CSV hasn't been loaded yet, or call load_csv() directly for a
    single file.

    Usage::

        with ElectionDatabase() as db:
            loader = LoadSummary(db)
            loader.sync(sources_dir=Path("sources"),
                        config_path=Path("elections.toml"))
    """

    def sync(
        self,
        sources_dir: Path = DEFAULT_SOURCES_DIR,
        config_path: Path = DEFAULT_CONFIG_PATH,
    ) -> dict[str, tuple[Election, list[str]]]:
        """
        Scan elections.toml for elections whose summary CSV hasn't been
        loaded yet and load them.

        Returns:
            Dict mapping summary filename → (Election, new_unrecognized_names)
            for each newly loaded file.
        """
        if not sources_dir.exists():
            raise FileNotFoundError(f"Sources directory not found: {sources_dir}")

        configs = load_elections_config(config_path)
        if not configs:
            print(f"  No elections found in {config_path}")
            return {}

        results: dict[str, tuple[Election, list[str]]] = {}
        for entry in configs:
            filename = entry["source_file"]
            if self._db.is_file_loaded(filename):
                continue

            # Election may already exist under this name. Register the CSV
            # as a known source so future syncs skip it, but don't re-insert.
            existing = self._db.get_election_by_name(entry["name"])
            if existing is not None:
                if existing.id is not None:
                    self._db.register_file(filename, existing.id)
                continue

            path = sources_dir / filename
            if not path.exists():
                print(f"  Skipping {filename}: file not found in {sources_dir}")
                continue

            election, new_names = self.load_csv(path, entry)
            results[filename] = (election, new_names)

        return results

    def load_csv(
        self,
        path: Path,
        config: dict,
    ) -> tuple[Election, list[str]]:
        """
        Load a single election CSV into the database.

        Args:
            path:   Path to the CSV file.
            config: Dict from elections.toml with at minimum 'name' and
                    'summary_file'.

        Returns:
            (Election, new_unrecognized_contest_names)
        """
        try:
            df = pd.read_csv(path, encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(path, encoding="windows-1252")
        df = _normalize_csv_columns(df)
        df = _validate_csv_columns(df, path)

        year = config.get("year") or _year_from_filename(path.name)
        if year is None:
            raise ValueError(
                f"Could not determine year for {path.name}. "
                "Add 'year' to elections.toml."
            )

        election = Election(
            id=None,
            name=config["name"],
            year=year,
            election_date=date.fromisoformat(config["election_date"])
            if config.get("election_date")
            else None,
            results_last_updated=date.fromisoformat(config["results_last_updated"])
            if config.get("results_last_updated")
            else None,
            summary_file=path.name,
            category=config.get("category", ""),
            election_type=config.get("election_type", ""),
            ballots_cast=config.get("ballots_cast"),
            registered_voters=config.get("registered_voters"),
        )

        election, new_names = self._db.insert_election(election, df)

        if election.id is None:
            raise RuntimeError("insert_election did not return an election id")
        self._db.register_file(path.name, election.id)
        return election, new_names


# ---------------------------------------------------------------------------
# Subclass 2: precinct-detail Excel loader
# ---------------------------------------------------------------------------


class LoadPrecinctDetail(_LoaderBase):
    """
    Reads the precinct-detail Excel workbook and populates
    candidate_precinct_results.

    The workbook has one sheet per party per contest. Each sheet contains:

      Row 0  – contest name (e.g. "FOR UNITED STATES SENATOR ( 1)")
      Row 1  – candidate names, one per block of 5 vote-type columns
               (Early, Vote by Mail, Polling, Provisional, Total Votes);
               col 0 is empty, col 1 is Registered Voters.
               "NO CANDIDATE/CANDIDATO" means the slot is uncontested.
      Row 2  – column headers: Precinct, Registered Voters,
               [Early, Vote by Mail, Polling, Provisional, Total Votes] × N,
               Total
      Row 3+ – one data row per precinct (last row is "Total:")

    All sheets for an election file are loaded into the same election_id.
    The detail filename is registered in loaded_files so re-running is safe.

    Usage::

        with ElectionDatabase() as db:
            loader = LoadPrecinctDetail(db)
            loader.sync(sources_dir=Path("sources"),
                        config_path=Path("elections.toml"))
    """

    # ---------------------------------------------------------------------------
    # Public interface
    # ---------------------------------------------------------------------------

    def sync(
        self,
        sources_dir: Path = DEFAULT_SOURCES_DIR,
        config_path: Path = DEFAULT_CONFIG_PATH,
    ) -> dict[str, tuple[Election, int]]:
        """
        Scan elections.toml for elections whose detail_file hasn't been
        loaded yet and load them.

        Returns:
            Dict mapping detail filename → (Election, rows_inserted)
            for each newly loaded file.
        """
        if not sources_dir.exists():
            raise FileNotFoundError(f"Sources directory not found: {sources_dir}")

        configs = load_elections_config(config_path)
        if not configs:
            return {}

        results: dict[str, tuple[Election, int]] = {}
        for entry in configs:
            detail_file = entry.get("detail_file")
            if not detail_file:
                continue
            if self._db.is_file_loaded(detail_file):
                continue

            election = self._db.get_election_by_name(entry["name"])
            if election is None:
                print(
                    f"  Skipping {detail_file}: election {entry['name']!r} "
                    "not yet in database. Run LoadSummary.sync() first."
                )
                continue

            path = sources_dir / detail_file
            if not path.exists():
                print(f"  Skipping {detail_file}: file not found in {sources_dir}")
                continue

            rows_inserted = self.load_detail_excel(path, election)
            results[detail_file] = (election, rows_inserted)

        return results

    def load_detail_excel(
        self,
        path: Path,
        election: Election,
    ) -> int:
        """
        Parse the detail Excel workbook at *path* and insert precinct-level
        results for *election*.

        Args:
            path:     Path to the .xlsx detail file.
            election: The Election this file belongs to. Must already have
                      a database id (i.e. the summary CSV was loaded first).

        Returns:
            Total number of rows inserted.

        Raises:
            ValueError  if election.id is None.
            FileNotFoundError if path does not exist.
        """
        if election.id is None:
            raise ValueError(
                f"Election {election.name!r} has no database id. "
                "Load the summary CSV first."
            )
        if not path.exists():
            raise FileNotFoundError(f"Detail file not found: {path}")

        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)

        # Build contest_name → contest_id lookup once for the whole workbook
        contest_id_map = self._build_contest_id_map()

        total_inserted = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            inserted = self._process_sheet(
                rows, election.id, contest_id_map, sheet_name
            )
            total_inserted += inserted

        self._db.register_file(path.name, election.id)
        return total_inserted

    # ---------------------------------------------------------------------------
    # Internal helpers
    # ---------------------------------------------------------------------------

    def _build_contest_id_map(self) -> dict[str, int]:
        """
        Return a mapping of normalized_contest_name → contest_id for every
        contest currently in the database.
        """
        df = self._db.query("SELECT id, contest_name FROM contests")
        return dict(zip(df["contest_name"], df["id"]))

    def _process_sheet(
        self,
        rows: list[tuple],
        election_id: int,
        contest_id_map: dict[str, int],
        sheet_name: str,
    ) -> int:
        """
        Parse one sheet and insert its precinct rows. Returns the number of
        rows inserted (0 for skipped/unrecognized sheets).
        """
        # Need at least: header row 0, candidate row 1, column row 2, 1 data row
        if len(rows) < 4:
            return 0

        # --- Row 0: contest name ---
        raw_contest = str(rows[0][0] or "").strip()
        if not raw_contest:
            return 0

        normalized_contest = normalize_contest_name(raw_contest)
        contest_id = contest_id_map.get(normalized_contest)
        if contest_id is None:
            # Contest not registered — sheet belongs to an election/category
            # outside the current database (e.g. a referendum not yet loaded).
            return 0

        # --- Row 1: candidate names ---
        # Layout: col 0=None, col 1=None, then 5-col blocks per candidate,
        # last col is "Total".
        # Candidate names appear at positions 2, 7, 12, … (every 5 columns).
        candidate_names: list[str] = []
        candidate_start_cols: list[int] = []
        candidate_row = rows[1]
        col = 2
        while col < len(candidate_row):
            cell = candidate_row[col]
            if cell is not None:
                name = str(cell).strip()
                if name.upper() in _NO_CANDIDATE_MARKERS:
                    col += 5
                    continue
                candidate_names.append(name)
                candidate_start_cols.append(col)
            col += 5

        if not candidate_names:
            return 0

        # --- Rows 3+: data rows ---
        # Skip the "Total:" summary row.
        precinct_rows_to_insert: list[dict] = []
        for data_row in rows[3:]:
            precinct_name = str(data_row[0] or "").strip()
            if not precinct_name or precinct_name.lower().startswith("total"):
                continue
            # col 1 = registered_voters (may be 0 for FEDERAL precincts)
            registered_voters_raw = data_row[1]
            registered_voters = (
                int(registered_voters_raw)
                if registered_voters_raw is not None
                else None
            )

            for i, (choice_name, start_col) in enumerate(
                zip(candidate_names, candidate_start_cols)
            ):
                # Each candidate block: Early, VBM, Polling, Provisional, Total
                try:
                    early = _int_or_zero(data_row[start_col])
                    vbm = _int_or_zero(data_row[start_col + 1])
                    polling = _int_or_zero(data_row[start_col + 2])
                    provisional = _int_or_zero(data_row[start_col + 3])
                    total_votes = _int_or_zero(data_row[start_col + 4])
                except IndexError:
                    continue

                precinct_rows_to_insert.append(
                    {
                        "election_id": election_id,
                        "contest_id": contest_id,
                        "contest_name_raw": raw_contest,
                        "choice_name": choice_name,
                        "precinct": precinct_name,
                        "registered_voters": registered_voters,
                        "early_votes": early,
                        "vote_by_mail": vbm,
                        "polling": polling,
                        "provisional": provisional,
                        "total_votes": total_votes,
                    }
                )

        if not precinct_rows_to_insert:
            return 0

        return self._db.insert_precinct_results(precinct_rows_to_insert)


def _int_or_zero(value) -> int:
    """Coerce a cell value to int, defaulting to 0 for None or non-numeric."""
    if value is None:
        return 0
    try:
        return int(value)
    except (ValueError, TypeError):
        return 0
