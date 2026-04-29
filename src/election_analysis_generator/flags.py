"""
election_analysis_generator/flags.py
-------------------------
All logic for exporting, importing, and interactively reviewing unresolved
contest name flags.

Three public functions mirror the three workflow scripts:

    export_flags(db, output_path)   – write flags_review.xlsx
    import_flags(db, input_path)    – read a reviewed workbook and apply changes
    review_flags(db)                – interactive CLI loop

The root-level scripts (export_flags.py, import_flags.py, review_flags.py)
are now thin shims that open the database and delegate here.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .db import ElectionDatabase

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DEFAULT_EXPORT_PATH = Path("flags_review.xlsx")
DEFAULT_IMPORT_PATH = Path("flags_review.xlsx")

VALID_STATUSES = frozenset({"accepted", "mapped", "ignored", "unreviewed"})


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------


def export_flags(
    db: ElectionDatabase,
    output_path: Path = DEFAULT_EXPORT_PATH,
) -> int:
    """Write unresolved flags to an Excel workbook for human review.

    The workbook has two sheets:
      - ``flags``         – one row per unresolved flag with a Status column
      - ``known_contests`` – all normalized contest names currently in the registry

    Returns the number of flags exported (0 if nothing to do).

    Workflow::

        with ElectionDatabase(db_path) as db:
            n = export_flags(db, Path("flags_review.xlsx"))
        # edit the workbook, then:
        with ElectionDatabase(db_path) as db:
            import_flags(db, Path("flags_review.xlsx"))
    """
    flags = db.get_unresolved_flags()
    known = sorted(db.get_known_contest_names())

    if not flags:
        return 0

    flags_df = pd.DataFrame(flags)[["id", "year", "contest_name_raw", "contest_name"]]
    flags_df = flags_df.rename(
        columns={
            "id": "Flag ID",
            "year": "Year",
            "contest_name_raw": "Raw Name",
            "contest_name": "Normalized Suggestion",
        }
    )
    flags_df["Status"] = "unreviewed"
    flags_df["Override Target"] = ""
    flags_df["Notes"] = ""

    known_df = pd.DataFrame({"Normalized Contest Name": known})

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        flags_df.to_excel(writer, sheet_name="flags", index=False)
        known_df.to_excel(writer, sheet_name="known_contests", index=False)
        _format_flags_sheet(writer.sheets["flags"], flags_df)
        _format_known_sheet(writer.sheets["known_contests"])

    return len(flags_df)


def _format_flags_sheet(ws, df: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    widths = {
        "Flag ID": 10,
        "Year": 8,
        "Raw Name": 55,
        "Normalized Suggestion": 55,
        "Status": 14,
        "Override Target": 55,
        "Notes": 35,
    }
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(
            col_name, 20
        )

    ws.freeze_panes = "A2"


def _format_known_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="375623")
    header_font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.column_dimensions["A"].width = 65
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------


def import_flags(
    db: ElectionDatabase,
    input_path: Path = DEFAULT_IMPORT_PATH,
) -> dict[str, int]:
    """Read a reviewed flags workbook and apply the decisions to the database.

    Processes rows where Status is one of:

    ``accepted``
        Register the Normalized Suggestion as a known contest name and mark
        the flag resolved.
    ``mapped``
        Add a raw-name → Override Target mapping, register the target, and
        mark the flag resolved.
    ``ignored``
        Mark the flag resolved without registering anything.
    ``unreviewed``
        Skip (no changes made).

    Returns a counts dict with keys ``accepted``, ``mapped``, ``ignored``,
    ``skipped``, and ``errors``.

    Raises ``FileNotFoundError`` if *input_path* does not exist.
    Raises ``ValueError`` if required columns are missing from the workbook.
    """
    if not input_path.exists():
        raise FileNotFoundError(f"Flags workbook not found: {input_path}")

    df = pd.read_excel(input_path, sheet_name="flags", dtype=str).fillna("")
    df.columns = [c.strip() for c in df.columns]

    required = {"Flag ID", "Year", "Raw Name", "Normalized Suggestion", "Status"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns in workbook: {missing}")

    df["Status"] = df["Status"].str.strip().str.lower()

    unrecognised = set(df["Status"].unique()) - VALID_STATUSES
    if unrecognised:
        # Warn but continue — those rows will fall through to 'skipped'.
        import warnings

        warnings.warn(
            f"Unrecognised Status values will be skipped: {unrecognised}",
            stacklevel=2,
        )

    counts: dict[str, int] = {
        "accepted": 0,
        "mapped": 0,
        "ignored": 0,
        "skipped": 0,
        "errors": 0,
    }

    for _, row in df.iterrows():
        status = row["Status"]
        flag_id = int(row["Flag ID"])
        year = int(row["Year"])
        raw_name = row["Raw Name"].strip()
        normalized = row["Normalized Suggestion"].strip()
        override_target = row.get("Override Target", "").strip()

        if status == "unreviewed":
            counts["skipped"] += 1

        elif status == "accepted":
            db.register_contest_name(normalized, year)
            db.resolve_flag(flag_id)
            counts["accepted"] += 1

        elif status == "mapped":
            known = db.get_known_contest_names()
            if not override_target:
                counts["errors"] += 1
                continue
            if override_target not in known:
                counts["errors"] += 1
                continue
            db.add_override(raw_name, override_target)
            db.register_contest_name(override_target, year)
            db.resolve_flag(flag_id)
            counts["mapped"] += 1

        elif status == "ignored":
            db.resolve_flag(flag_id)
            counts["ignored"] += 1

        else:
            counts["skipped"] += 1

    return counts


# ---------------------------------------------------------------------------
# Interactive review
# ---------------------------------------------------------------------------


def review_flags(db: ElectionDatabase) -> None:
    """Interactive CLI loop for resolving flags one at a time.

    For each unresolved flag, the user chooses:

    ``[a]`` Accept as a new contest name.
    ``[m]`` Map to an existing contest name (searchable).
    ``[s]`` Skip for now.
    """
    flags = db.get_unresolved_flags()
    if not flags:
        print("No unresolved contest name flags.")
        return

    print(f"{len(flags)} unresolved flag(s).\n")

    for flag in flags:
        flag_id = flag["id"]
        year = flag["year"]
        raw_name = flag["contest_name_raw"]
        norm = flag["contest_name"]

        print(f"Year:       {year}")
        print(f"Raw name:   {raw_name}")
        print(f"Normalized: {norm}")
        print()
        print("  [a] Accept as new contest")
        print("  [m] Map to an existing contest name")
        print("  [s] Skip")
        print()

        while True:
            choice = input("Choice: ").strip().lower()

            if choice == "a":
                db.register_contest_name(norm, year)
                db.resolve_flag(flag_id)
                print("  ✓ Accepted.\n")
                break

            elif choice == "m":
                known = sorted(db.get_known_contest_names())
                query = (
                    input("  Search existing names (or Enter to list all): ")
                    .strip()
                    .lower()
                )
                matches = [n for n in known if query in n.lower()] if query else known

                if not matches:
                    print("  No matches found.")
                    continue

                for i, name in enumerate(matches[:20], 1):
                    print(f"  {i:>2}. {name}")
                if len(matches) > 20:
                    print(f"  ... and {len(matches) - 20} more. Refine your search.")
                    continue

                idx = input("  Enter number to select: ").strip()
                if not idx.isdigit() or not (1 <= int(idx) <= len(matches)):
                    print("  Invalid selection.")
                    continue

                canonical = matches[int(idx) - 1]
                note = input(f"  Note (optional, e.g. 'Renamed in {year}'): ").strip()
                db.add_override(raw_name, canonical, note or None)
                db.resolve_flag(flag_id)
                print(f"  ✓ Mapped to: {canonical}\n")
                break

            elif choice == "s":
                print("  Skipped.\n")
                break

            else:
                print("  Please enter a, m, or s.")

    remaining = len(db.get_unresolved_flags())
    if remaining:
        print(f"{remaining} flag(s) still unresolved. Run again to continue.")
    else:
        print("All flags resolved.")
