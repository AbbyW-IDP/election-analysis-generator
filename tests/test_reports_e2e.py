"""
test_reports_e2e.py
-------------------
End-to-end tests for report generation.

These tests exercise the full pipeline:
  load_reports_config() -> LoadSummary.sync() -> LoadPrecinctDetail.sync()
  -> run_reports() -> Excel output on disk

Fixtures live in tests/fixtures/. The DB is loaded once per session
(session scope) because parsing the real CSVs and Excel detail file is
slow enough to matter if repeated per test.

Contest name flags are left unresolved (Option C): the report pipeline
reads from contest_results directly and does not filter on flag status,
so unresolved flags do not affect analysis output.

Some tests are marked xfail where a known issue in the codebase is
expected to cause a failure. These will be removed once the underlying
bug is fixed.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.election_analysis_generator.db import ElectionDatabase
from src.election_analysis_generator.loader import LoadSummary, LoadPrecinctDetail
from src.election_analysis_generator.reports import load_reports_config, run_reports

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

FIXTURES_DIR = Path("tests/fixtures")
SOURCES_DIR = FIXTURES_DIR / "sources"
ELECTIONS_CSV = FIXTURES_DIR / "elections.csv"
REPORTS_TOML = FIXTURES_DIR / "reports.toml"

# ---------------------------------------------------------------------------
# Session-scoped fixtures: load data once, run reports once
# ---------------------------------------------------------------------------


@pytest.fixture(scope="session")
def e2e_db():
    """In-memory DB loaded from real fixture CSVs and the precinct detail Excel.

    Contest name flags are left unresolved -- the report pipeline does not
    filter on flag status so this does not affect analysis output.
    """
    db = ElectionDatabase(db_path=":memory:")
    LoadSummary(db).sync(sources_dir=SOURCES_DIR, config_path=ELECTIONS_CSV)
    LoadPrecinctDetail(db).sync(sources_dir=SOURCES_DIR, config_path=ELECTIONS_CSV)
    yield db
    db.close()


@pytest.fixture(scope="session")
def e2e_report_paths(e2e_db, tmp_path_factory):
    """Run all reports defined in the fixture reports.toml once per session.

    Returns the list of Path objects written by run_reports().
    """
    # Arrange
    out_dir = tmp_path_factory.mktemp("e2e_reports")
    configs = load_reports_config(REPORTS_TOML)

    # Act
    paths = run_reports(configs, e2e_db, base_dir=out_dir)
    return paths


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------


def _read_sheet(paths: list[Path], sheet: str) -> pd.DataFrame:
    """Read a named sheet from the first (and only) output workbook."""
    return pd.read_excel(paths[0], sheet_name=sheet)


# ---------------------------------------------------------------------------
# Test: output files
# ---------------------------------------------------------------------------


def test_run_reports_writes_one_file(e2e_report_paths: list[Path]) -> None:
    """run_reports() writes exactly one Excel file for the one configured report."""
    # Arrange / Act: done in session fixture

    # Assert
    assert len(e2e_report_paths) == 1
    assert e2e_report_paths[0].exists()
    assert e2e_report_paths[0].suffix == ".xlsx"


# ---------------------------------------------------------------------------
# Test: sheet existence
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "sheet_name",
    [
        pytest.param("turnout", id="turnout"),
        pytest.param("22-26 pct change by party", id="pct_change_by_party"),
        pytest.param("22-26 party share", id="party_share"),
        pytest.param("raw data", id="aggregated_csv"),
        pytest.param("precinct turnout", id="precinct_turnout"),
    ],
)
def test_sheet_exists(e2e_report_paths: list[Path], sheet_name: str) -> None:
    """Every analysis entry in reports.toml produces a sheet in the workbook."""
    # Arrange
    wb = load_workbook(e2e_report_paths[0], read_only=True)

    # Act / Assert
    assert sheet_name in wb.sheetnames, (
        f"Sheet {sheet_name!r} missing. Found: {wb.sheetnames}"
    )


# ---------------------------------------------------------------------------
# Test: sheet shape -- columns and minimum row count
# ---------------------------------------------------------------------------


def test_turnout_shape(e2e_report_paths: list[Path]) -> None:
    """turnout sheet has one column per election and the three metric rows."""
    # Arrange
    # turnout() is written with index=True, so the first column is the
    # row index (Metric). Use index_col=0 so the metric labels become the
    # index rather than an unnamed column.
    df = pd.read_excel(e2e_report_paths[0], sheet_name="turnout", index_col=0)

    # Act / Assert: one column per election
    assert "2022 General Primary" in df.columns
    assert "2026 General Primary" in df.columns

    # Three metric rows: % Vote, Registered, Ballots Cast
    assert len(df) == 3


def test_pct_change_by_party_shape(e2e_report_paths: list[Path]) -> None:
    """pct_change_by_party sheet has the expected column set and at least one row."""
    # Arrange
    df = _read_sheet(e2e_report_paths, "22-26 pct change by party")
    expected_cols = [
        "contest",
        "DEM 2022 General Primary",
        "DEM 2026 General Primary",
        "DEM % change",
        "REP 2022 General Primary",
        "REP 2026 General Primary",
        "REP % change",
    ]

    # Act / Assert
    for col in expected_cols:
        assert col in df.columns, (
            f"missing column {col!r} in pct_change_by_party sheet"
        )
    assert len(df) >= 1


def test_party_share_shape(e2e_report_paths: list[Path]) -> None:
    """party_share sheet has the expected column set and at least one row."""
    # Arrange
    df = _read_sheet(e2e_report_paths, "22-26 party share")
    expected_cols = [
        "contest",
        "DEM share 2022 General Primary",
        "DEM share 2026 General Primary",
        "DEM pp change",
        "REP share 2022 General Primary",
        "REP share 2026 General Primary",
        "REP pp change",
    ]

    # Act / Assert
    for col in expected_cols:
        assert col in df.columns, (
            f"missing column {col!r} in party_share sheet"
        )
    assert len(df) >= 1


def test_aggregated_csv_shape(e2e_report_paths: list[Path]) -> None:
    """aggregated_csv sheet has the expected columns and one row per candidate."""
    # Arrange
    df = _read_sheet(e2e_report_paths, "raw data")
    expected_cols = [
        "contest name",
        "choice name",
        "party",
        "total votes",
        "year",
        "election name",
        "contest name (normalized)",
        "category",
    ]

    # Act / Assert
    for col in expected_cols:
        assert col in df.columns, (
            f"missing column {col!r} in aggregated_csv sheet"
        )
    assert len(df) >= 1


def test_precinct_turnout_shape(e2e_report_paths: list[Path]) -> None:
    """precinct_turnout sheet has the expected columns and at least one data row."""
    # Arrange
    df = _read_sheet(e2e_report_paths, "precinct turnout")
    expected_cols = [
        "election",
        "year",
        "contest",
        "party",
        "candidate",
        "precinct",
        "registered_voters",
        "total_votes",
        "turnout_rate",
    ]

    # Act / Assert
    for col in expected_cols:
        assert col in df.columns, (
            f"missing column {col!r} in precinct_turnout sheet"
        )
    assert len(df) >= 1
